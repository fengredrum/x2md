"""Excel (.xlsx/.xls) to Markdown conversion with HTML table output.

Parses workbook worksheets and generates Markdown with HTML <table> elements
preserving merged cells (colspan/rowspan) and cell styles (bold, italic, font
size, colors, background, alignment, borders).
"""

from __future__ import annotations

from pathlib import Path


def _escape_html(text: str) -> str:
    """Escape special HTML characters in cell text content."""
    text = (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br>")
    )
    return text


def _color_to_css(color_obj) -> str | None:
    """Convert an openpyxl Color object to a CSS hex color string.

    Handles:
    - RGB colors: ``'005196FD'`` (8-char ARGB) → ``'#5196FD'``
    - Indexed colors: looks up in openpyxl's COLOR_INDEX palette
    - Theme colors: not currently resolved (returns None)
    - None → None
    """
    if color_obj is None:
        return None

    # Try raw rgb string first (both 6-char and 8-char ARGB)
    rgb_val = getattr(color_obj, "rgb", None)
    if isinstance(rgb_val, str) and len(rgb_val) >= 6:
        if len(rgb_val) == 8:
            rgb_val = rgb_val[2:]  # strip ARGB alpha prefix
        if len(rgb_val) == 6:
            return f"#{rgb_val}"

    # Fall back to indexed color lookup
    indexed = getattr(color_obj, "indexed", None)
    if isinstance(indexed, int):
        try:
            from openpyxl.styles.colors import COLOR_INDEX

            argb = COLOR_INDEX[indexed]
            if isinstance(argb, str) and len(argb) == 8:
                return f"#{argb[2:]}"
        except (IndexError, ImportError):
            pass

    return None


def _border_name_to_css(style_name: str | None) -> str | None:
    """Map an openpyxl border style name to a CSS border style."""
    if style_name is None:
        return None
    mapping = {
        "thin": "1px solid",
        "medium": "2px solid",
        "thick": "3px solid",
        "double": "3px double",
        "dashed": "1px dashed",
        "dotted": "1px dotted",
        "hair": "1px solid",
        "mediumDashed": "2px dashed",
        "dashDot": "1px dashed",
        "mediumDashDot": "2px dashed",
        "dashDotDot": "1px dotted",
        "mediumDashDotDot": "2px dotted",
        "slantDashDot": "1px dashed",
    }
    return mapping.get(style_name)


def _cell_style_openpyxl(cell) -> str:
    """Build an inline CSS ``style`` attribute value from an openpyxl cell.

    Extracts font, fill, alignment, and border properties and converts them
    to a semicolon-separated CSS string.
    """
    styles: list[str] = []

    # ── Font ──
    font = cell.font
    if font:
        if font.bold:
            styles.append("font-weight: bold")
        if font.italic:
            styles.append("font-style: italic")
        if font.size:
            styles.append(f"font-size: {font.size}pt")
        css_color = _color_to_css(font.color)
        if css_color:
            styles.append(f"color: {css_color}")

    # ── Fill (background) ──
    fill = cell.fill
    if fill:
        css_bg = _color_to_css(fill.fgColor)
        if css_bg:
            styles.append(f"background-color: {css_bg}")

    # ── Alignment ──
    align = cell.alignment
    if align:
        if align.horizontal and align.horizontal != "general":
            styles.append(f"text-align: {align.horizontal}")
        if align.vertical and align.vertical != "bottom":
            styles.append(f"vertical-align: {align.vertical}")

    # ── Borders ──
    border = cell.border
    if border:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name, None)
            if side and side.style:
                css_border = _border_name_to_css(side.style)
                if css_border:
                    color_part = ""
                    side_css = _color_to_css(side.color)
                    if side_css:
                        color_part = f" {side_css}"
                    styles.append(
                        f"border-{side_name}: {css_border}{color_part}"
                    )

    return "; ".join(styles)


def _cell_value_to_str(value) -> str:
    """Convert a cell value to an HTML-safe string.

    Handles None, str, int, float, datetime, and bool types.
    Newlines become ``<br>`` and special HTML chars are escaped.
    """
    if value is None:
        return ""

    s = str(value).strip()

    # Escape HTML entities
    s = _escape_html(s)

    return s


def _worksheet_to_html(ws) -> str:
    """Convert an openpyxl worksheet to an HTML ``<table>`` string.

    Preserves merged cells (colspan/rowspan) and cell styles (inline CSS).
    Header rows (row 1 or bold cells) use ``<th>``; others use ``<td>``.
    """
    if ws.max_row is None or ws.max_column is None:
        return ""

    max_row = ws.max_row
    max_col = ws.max_column

    # ── Pass 1: Build merged cell maps ──
    # merged: (row, col) → (rowspan, colspan) for the anchor cell
    # skip:   (row, col) → True for cells covered by a span
    merged: dict[tuple[int, int], tuple[int, int]] = {}
    skip: set[tuple[int, int]] = set()

    for merge_range in ws.merged_cells.ranges:
        rspan = merge_range.max_row - merge_range.min_row + 1
        cspan = merge_range.max_col - merge_range.min_col + 1
        anchor = (merge_range.min_row, merge_range.min_col)
        merged[anchor] = (rspan, cspan)
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(
                merge_range.min_col, merge_range.max_col + 1
            ):
                if (r, c) != anchor:
                    skip.add((r, c))

    # ── Pass 2: Generate HTML ──
    html_parts: list[str] = ["<table>"]

    # Build a quick lookup of cell objects by coordinate
    # We iterate ws rows, which is more efficient than random access
    cell_cache: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            cell_cache[(cell.row, cell.column)] = cell

    for row_idx in range(1, max_row + 1):
        # Check if all cells in this row are covered by rowspan from above.
        # If so, emit an empty <tr></tr> to preserve rowspan offsets.
        all_in_skip = True
        any_content = False
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) not in skip:
                all_in_skip = False
                cell = cell_cache.get((row_idx, col_idx))
                if cell is not None and cell.value is not None:
                    any_content = True

        if all_in_skip:
            # Entire row is consumed by rowspan from a prior merged cell.
            # Emit empty row so the rowspan count remains correct.
            html_parts.append("  <tr></tr>")
            continue

        if not any_content:
            continue

        html_parts.append("  <tr>")

        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in skip:
                continue

            cell = cell_cache.get((row_idx, col_idx))

            # Build attributes
            attrs: list[str] = []

            # Style
            if cell is not None:
                style_str = _cell_style_openpyxl(cell)
                if style_str:
                    attrs.append(f'style="{style_str}"')

            # Colspan / rowspan
            span = merged.get((row_idx, col_idx))
            if span:
                rspan, cspan = span
                if rspan > 1:
                    attrs.append(f'rowspan="{rspan}"')
                if cspan > 1:
                    attrs.append(f'colspan="{cspan}"')

            # <th> vs <td>
            is_header = (row_idx == 1) or (
                cell is not None
                and cell.font
                and cell.font.bold
            )
            tag = "th" if is_header else "td"

            # Cell value
            value_str = ""
            if cell is not None:
                value_str = _cell_value_to_str(cell.value)

            attr_str = (" " + " ".join(attrs)) if attrs else ""
            html_parts.append(f"    <{tag}{attr_str}>{value_str}</{tag}>")

        html_parts.append("  </tr>")

    html_parts.append("</table>")
    return "\n".join(html_parts)


def _convert_xlsx(input_path: Path, output_path: Path) -> None:
    """Convert a .xlsx file to Markdown using openpyxl.

    Each worksheet becomes an ``##`` section with an HTML table.
    """
    import openpyxl

    wb = openpyxl.load_workbook(input_path, data_only=True)

    parts: list[str] = []
    parts.append(f"# {input_path.stem}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## {sheet_name}\n")

        html_table = _worksheet_to_html(ws)
        if html_table:
            parts.append(html_table)
        else:
            parts.append("_（空表）_\n")
        parts.append("")  # blank line between sheets

    wb.close()
    output_path.write_text("\n".join(parts), encoding="utf-8")


def _convert_xls(input_path: Path, output_path: Path) -> None:
    """Convert a .xls file to Markdown using xlrd.

    Each worksheet becomes an ``##`` section with an HTML table.
    Merged cells are preserved via colspan/rowspan.
    """
    try:
        import xlrd
    except ImportError:
        raise ImportError(
            "xlrd is required for .xls (legacy Excel) conversion. "
            "Install with: pip install x2md[xls]"
        )

    wb = xlrd.open_workbook(str(input_path), formatting_info=True)

    parts: list[str] = []
    parts.append(f"# {input_path.stem}\n")

    for sheet_idx in range(wb.nsheets):
        sheet = wb.sheet_by_index(sheet_idx)
        sheet_name = sheet.name
        parts.append(f"## {sheet_name}\n")

        if sheet.nrows == 0:
            parts.append("_（空表）_\n")
            parts.append("")
            continue

        # ── Build merged cell maps ──
        # xlrd: (rlo, rhi, clo, chi) — 0-based, end-exclusive
        merged: dict[tuple[int, int], tuple[int, int]] = {}
        skip: set[tuple[int, int]] = set()

        for rlo, rhi, clo, chi in sheet.merged_cells:
            rspan = rhi - rlo
            cspan = chi - clo
            anchor = (rlo, clo)  # 0-based
            merged[anchor] = (rspan, cspan)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    if (r, c) != anchor:
                        skip.add((r, c))

        # ── Generate HTML ──
        html_parts: list[str] = ["<table>"]

        for row_idx in range(sheet.nrows):
            # Check if row has content
            row_has_content = any(
                sheet.cell_value(row_idx, col_idx) != ""
                for col_idx in range(sheet.ncols)
                if (row_idx, col_idx) not in skip
            )
            if not row_has_content:
                continue

            html_parts.append("  <tr>")

            for col_idx in range(sheet.ncols):
                if (row_idx, col_idx) in skip:
                    continue

                cell_value = sheet.cell_value(row_idx, col_idx)

                # Build attributes
                attrs: list[str] = []

                # Colspan / rowspan
                span = merged.get((row_idx, col_idx))
                if span:
                    rspan, cspan = span
                    if rspan > 1:
                        attrs.append(f'rowspan="{rspan}"')
                    if cspan > 1:
                        attrs.append(f'colspan="{cspan}"')

                # <th> vs <td> — only row 0 for xlrd (limited style info)
                tag = "th" if row_idx == 0 else "td"

                value_str = _escape_html(str(cell_value).strip()) if cell_value else ""

                attr_str = (" " + " ".join(attrs)) if attrs else ""
                html_parts.append(
                    f"    <{tag}{attr_str}>{value_str}</{tag}>"
                )

            html_parts.append("  </tr>")

        html_parts.append("</table>")
        parts.append("\n".join(html_parts))
        parts.append("")

    output_path.write_text("\n".join(parts), encoding="utf-8")


def convert_excel(
    input_path: str | Path,
    output_dir: str | Path | None = None,
) -> Path:
    """Convert an Excel file (.xlsx/.xls) to Markdown with HTML tables.

    Each worksheet becomes a ``##`` section. Tables are converted to HTML
    format preserving merged cells (colspan/rowspan) and cell styles
    (bold, italic, font size, colors, background, alignment, borders).

    .xlsx files are processed with openpyxl (full style support).
    .xls files are processed with xlrd (limited style support).

    Args:
        input_path: Path to the input .xlsx or .xls file.
        output_dir: Output directory (default: same directory as input).

    Returns:
        Path to the generated .md file.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If the input file is not a .xlsx or .xls file.
        ImportError: If xlrd is not installed for .xls conversion.
    """
    input_path = Path(input_path).resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    suffix = input_path.suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise ValueError(
            f"Unsupported format: {suffix}. Expected .xlsx or .xls"
        )

    output_dir = Path(output_dir) if output_dir else input_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    output_name = input_path.stem + ".md"
    output_path = output_dir / output_name

    if suffix == ".xlsx":
        _convert_xlsx(input_path, output_path)
    else:
        _convert_xls(input_path, output_path)

    return output_path
